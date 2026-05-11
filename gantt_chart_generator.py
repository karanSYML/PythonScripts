#!/usr/bin/env python3
"""
Gantt chart generator for CTO presentation.
Output: gantt_chart_2026_2027.xlsx

Three workstreams:
  1. Compatibility Analysis of Endurance  (Jun 2026 – Mar 2027)
  2. System Simulator Development         (Jun 2026 – Aug 2026)
  3. Mission Analysis Support             (Jun 2026 – May 2027)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "gantt_chart_2026_2027.xlsx"

MONTHS = [
    "Jun '26", "Jul '26", "Aug '26", "Sep '26", "Oct '26", "Nov '26",
    "Dec '26", "Jan '27", "Feb '27", "Mar '27", "Apr '27", "May '27",
]
N = len(MONTHS)  # 12

# Column layout (1-based)
COL_TASK  = 1   # A – task name
COL_START = 2   # B – start month label
COL_END   = 3   # C – end month label
COL_M0    = 4   # D – first month (Jun '26);  last = COL_M0 + N - 1 = 15 (O)

# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────
def mk_fill(hex6):
    return PatternFill("solid", fgColor=hex6)

def mk_font(hex6="000000", bold=False, sz=10, italic=False):
    return Font(color=hex6, bold=bold, size=sz, italic=italic, name="Calibri")

def mk_border(hex6="D9D9D9"):
    s = Side(style="thin", color=hex6)
    return Border(left=s, right=s, top=s, bottom=s)

def mk_align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

CTR   = mk_align("center")
LEFT1 = mk_align("left", indent=1)
LEFT3 = mk_align("left", indent=3)

# ─────────────────────────────────────────────────────────────────────────────
# Color palette  (hex without #)
# ─────────────────────────────────────────────────────────────────────────────
#   compat  → Blue family
#   sim     → Orange family
#   mission → Green family
PALETTE = {
    "title_bg":       mk_fill("1F4E79"),
    "hdr_dark":       mk_fill("2F2F2F"),
    "hdr_mid":        mk_fill("505050"),

    # Compatibility – Blue
    "compat_cat":     mk_fill("1F4E79"),   # dark navy
    "compat_cat_bar": mk_fill("2E75B6"),   # medium blue
    "compat_bar":     mk_fill("9DC3E6"),   # light blue bar
    "compat_row":     mk_fill("EBF3FB"),   # very light blue row bg
    "compat_cfont":   mk_font("FFFFFF", bold=True, sz=11),
    "compat_sfont":   mk_font("1A3A5C", sz=10),

    # System Simulator – Orange
    "sim_cat":        mk_fill("843C0C"),   # dark burnt orange
    "sim_cat_bar":    mk_fill("ED7D31"),   # orange
    "sim_bar":        mk_fill("F4B183"),   # light orange bar
    "sim_row":        mk_fill("FEF3E9"),   # very light orange row bg
    "sim_cfont":      mk_font("FFFFFF", bold=True, sz=11),
    "sim_sfont":      mk_font("5C2A06", sz=10),

    # Mission Analysis – Green
    "mission_cat":    mk_fill("1E3D1A"),   # dark forest green
    "mission_cat_bar":mk_fill("548235"),   # medium green
    "mission_bar":    mk_fill("A9D18E"),   # light green bar
    "mission_row":    mk_fill("E2F0D9"),   # very light green row bg
    "mission_cfont":  mk_font("FFFFFF", bold=True, sz=11),
    "mission_sfont":  mk_font("1A3D12", sz=10),

    # Grid
    "month_even":     mk_fill("F7F7F7"),
    "month_odd":      mk_fill("FFFFFF"),
    "spacer":         mk_fill("FFFFFF"),
}

# ─────────────────────────────────────────────────────────────────────────────
# Task data
#   (name, start_idx, end_idx, level, group)
#   level  0 = category header row
#          1 = subtask row
#         -1 = spacer row (name/group ignored)
#   Months: 0=Jun'26 … 11=May'27
# ─────────────────────────────────────────────────────────────────────────────
TASKS = [
    # ── 1. Compatibility Analysis of Endurance ─────────────────────────────
    ("Compatibility Analysis of Endurance",  0,  9, 0, "compat"),
    ("Requirements Review",                  0,  0, 1, "compat"),   # Jun
    ("Interface Analysis",                   1,  2, 1, "compat"),   # Jul–Aug
    ("Structural / Thermal Compatibility",   2,  4, 1, "compat"),   # Aug–Oct
    ("Electrical Compatibility",             4,  6, 1, "compat"),   # Oct–Dec
    ("SW Integration",                       5,  7, 1, "compat"),   # Nov–Jan
    ("Verification & Validation",            7,  8, 1, "compat"),   # Jan–Feb
    ("Final Report & Closure",               8,  9, 1, "compat"),   # Feb–Mar

    ("", 0, 0, -1, None),   # spacer

    # ── 2. System Simulator Development ────────────────────────────────────
    ("System Simulator Development",         0,  2, 0, "sim"),
    ("Design & Architecture",                0,  0, 1, "sim"),      # Jun
    ("Development & Integration",            0,  1, 1, "sim"),      # Jun–Jul
    ("Testing & Delivery",                   2,  2, 1, "sim"),      # Aug

    ("", 0, 0, -1, None),   # spacer

    # ── 3. Mission Analysis Support ─────────────────────────────────────────
    ("Mission Analysis Support",             0, 11, 0, "mission"),
    ("Sloshing Analysis",                    0,  3, 1, "mission"),  # Jun–Sep
    ("Plume Impingement",                    1,  4, 1, "mission"),  # Jul–Oct
    ("Venting & Contamination",              2,  5, 1, "mission"),  # Aug–Nov
    ("Radiation Analysis",                   3,  6, 1, "mission"),  # Sep–Dec
    ("ESD Analysis",                         4,  7, 1, "mission"),  # Oct–Jan
    ("EMC Analysis",                         5,  8, 1, "mission"),  # Nov–Feb
    ("HFSS Simulation",                      6,  9, 1, "mission"),  # Dec–Mar
    ("ITU Simulation",                       8, 11, 1, "mission"),  # Feb–May
]


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"
    ws.sheet_view.showGridLines = False

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 34
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1,   end_column=COL_M0 + N - 1,
    )
    t = ws.cell(row=1, column=1,
                value="Activity Plan  ·  Karan Anand  ·  Jun 2026 – May 2027")
    t.fill      = PALETTE["title_bg"]
    t.font      = mk_font("FFFFFF", bold=True, sz=14)
    t.alignment = CTR

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[2].height = 22
    for col, lbl in [(COL_TASK, "Task / Deliverable"),
                     (COL_START, "Start"),
                     (COL_END, "End")]:
        c = ws.cell(row=2, column=col, value=lbl)
        c.fill      = PALETTE["hdr_dark"]
        c.font      = mk_font("FFFFFF", bold=True, sz=10)
        c.alignment = CTR
        c.border    = mk_border("444444")

    for i, m in enumerate(MONTHS):
        c = ws.cell(row=2, column=COL_M0 + i, value=m)
        c.fill      = PALETTE["hdr_dark"] if i % 2 == 0 else PALETTE["hdr_mid"]
        c.font      = mk_font("FFFFFF", bold=True, sz=9)
        c.alignment = CTR
        c.border    = mk_border("444444")

    # ── Task rows ─────────────────────────────────────────────────────────────
    START_ROW = 3
    subtask_rows = []   # collect for outline grouping

    for offset, (name, s, e, level, grp) in enumerate(TASKS):
        r = START_ROW + offset

        # ── Spacer ──
        if level == -1:
            ws.row_dimensions[r].height = 6
            for col in range(1, COL_M0 + N):
                c = ws.cell(row=r, column=col)
                c.fill = PALETTE["spacer"]
            continue

        cat_fill  = PALETTE[f"{grp}_cat"]
        cat_cfont = PALETTE[f"{grp}_cfont"]
        cat_bar   = PALETTE[f"{grp}_cat_bar"]
        sub_fill  = PALETTE[f"{grp}_row"]
        sub_sfont = PALETTE[f"{grp}_sfont"]
        bar_fill  = PALETTE[f"{grp}_bar"]

        if level == 0:
            # ── Category header ──────────────────────────────────────────────
            ws.row_dimensions[r].height = 22

            c = ws.cell(row=r, column=COL_TASK, value=name.upper())
            c.fill      = cat_fill
            c.font      = cat_cfont
            c.alignment = LEFT1
            c.border    = mk_border("666666")

            for col, lbl in [(COL_START, MONTHS[s]), (COL_END, MONTHS[e])]:
                cc = ws.cell(row=r, column=col, value=lbl)
                cc.fill      = cat_fill
                cc.font      = mk_font("DDDDDD", sz=9, italic=True)
                cc.alignment = CTR
                cc.border    = mk_border("666666")

            for i in range(N):
                cc = ws.cell(row=r, column=COL_M0 + i)
                cc.fill   = cat_bar if s <= i <= e else cat_fill
                cc.border = mk_border("555555")

        else:
            # ── Subtask ──────────────────────────────────────────────────────
            ws.row_dimensions[r].height = 17
            subtask_rows.append(r)

            c = ws.cell(row=r, column=COL_TASK, value=name)
            c.fill      = sub_fill
            c.font      = sub_sfont
            c.alignment = LEFT3
            c.border    = mk_border("CCCCCC")

            for col, lbl in [(COL_START, MONTHS[s]), (COL_END, MONTHS[e])]:
                cc = ws.cell(row=r, column=col, value=lbl)
                cc.fill      = sub_fill
                cc.font      = mk_font("555555", sz=9, italic=True)
                cc.alignment = CTR
                cc.border    = mk_border("CCCCCC")

            for i in range(N):
                cc = ws.cell(row=r, column=COL_M0 + i)
                if s <= i <= e:
                    cc.fill = bar_fill
                else:
                    cc.fill = PALETTE["month_even"] if i % 2 == 0 else PALETTE["month_odd"]
                cc.border = mk_border("CCCCCC")

    # ── Row outlining so subtasks can be collapsed ────────────────────────────
    for r in subtask_rows:
        ws.row_dimensions[r].outline_level = 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(COL_TASK)].width  = 36
    ws.column_dimensions[get_column_letter(COL_START)].width = 11
    ws.column_dimensions[get_column_letter(COL_END)].width   = 11
    for i in range(N):
        ws.column_dimensions[get_column_letter(COL_M0 + i)].width = 8

    # ── Freeze: headers + task-name column stay visible while scrolling ───────
    ws.freeze_panes = ws.cell(row=START_ROW, column=COL_M0)

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation  = "landscape"
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.print_title_rows        = "1:2"

    wb.save(OUTPUT_FILE)
    print(f"Saved → {OUTPUT_FILE}")


if __name__ == "__main__":
    build()
